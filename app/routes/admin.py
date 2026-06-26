from flask import Blueprint, render_template, redirect, url_for, request, flash
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from app.models import Admin, Kategori, Lokasi, ActivityLog  # Ditambahkan: ActivityLog
from app import db, mail
import os
import shutil  
import secrets 
from datetime import datetime, timedelta 
from flask_mail import Message 
from werkzeug.utils import secure_filename
from flask import current_app
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from flask import send_file

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']

admin_bp = Blueprint('admin', __name__)


# ─── FUNGSI HELPER: LOG AKTIVITAS (SINKRON DENGAN USER_ID) ────────────
def log_aktivitas(aksi, deskripsi):
    """
    Helper otomatis untuk menyimpan riwayat audit log berdasarkan current_user.id
    """
    try:
        id_admin = current_user.id if current_user.is_authenticated else None
        log = ActivityLog(user_id=id_admin, aksi=aksi, deskripsi=deskripsi)
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Gagal mencatat audit log: {str(e)}")


# ─── LOGIN ───────────────────────────────────────────
@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('admin.dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = Admin.query.filter_by(username=username).first()

        if user and check_password_hash(user.password, password):
            login_user(user)
            log_aktivitas('login', 'berhasil masuk ke sistem dashboard admin')
            return redirect(url_for('admin.dashboard'))
        else:
            flash('Username atau password salah.', 'danger')

    return render_template('admin/login.html')


# ─── LUPA PASSWORD ───────────────────────────────────
@admin_bp.route('/lupa-password', methods=['GET', 'POST'])
def lupa_password():
    if current_user.is_authenticated:
        return redirect(url_for('admin.dashboard'))
        
    if request.method == 'POST':
        email_input = request.form.get('email')
        user = Admin.query.filter_by(email=email_input).first()
        
        if user:
            token_acak = secrets.token_hex(16)
            user.reset_token = token_acak
            user.token_expiry = datetime.now() + timedelta(minutes=30)
            db.session.commit()
            
            link_reset = url_for('admin.reset_password', token=token_acak, _external=True)
            
            msg = Message("Reset Password WebGIS Karang Harapan", recipients=[user.email])
            msg.body = f"""Halo Admin Kelurahan Karang Harapan,

Kami menerima permintaan untuk menyetel ulang kata sandi akun WebGIS Anda.
Silakan klik tautan di bawah ini untuk mengganti password Anda:

{link_reset}

Catatan: Tautan ini hanya berlaku selama 30 menit demi alasan keamanan.
Jika Anda tidak merasa melakukan tindakan ini, abaikan email ini secara aman.

Salam hangat,
Tim Teknis WebGIS Kelurahan Karang Harapan
"""
            try:
                mail.send(msg)
                flash('Link reset password berhasil dikirim! Silakan periksa kotak masuk email desa Anda.', 'success')
                return redirect(url_for('admin.login'))
            except Exception as e:
                flash(f'Gagal mengirim email secara otomatis: {str(e)}', 'danger')
        else:
            flash('Alamat email tersebut tidak terdaftar di sistem admin kami.', 'danger')
            
    return render_template('admin/lupa_password.html')


# ─── RESET PASSWORD ──────────────────────────────────
@admin_bp.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if current_user.is_authenticated:
        return redirect(url_for('admin.dashboard'))
        
    user = Admin.query.filter(Admin.reset_token == token, Admin.token_expiry > datetime.now()).first()
    
    if not user:
        flash('Token reset password tidak valid atau sudah kedaluwarsa. Silakan ajukan ulang.', 'danger')
        return redirect(url_for('admin.lupa_password'))
        
    if request.method == 'POST':
        password_baru = request.form.get('password_baru')
        konfirmasi = request.form.get('konfirmasi')
        
        if len(password_baru) < 6:
            flash('Password baru minimal berjumlah 6 karakter.', 'danger')
        elif password_baru != konfirmasi:
            flash('Konfirmasi password baru tidak cocok.', 'danger')
        else:
            user.password = generate_password_hash(password_baru)
            user.reset_token = None
            user.token_expiry = None
            db.session.commit()
            
            flash('Password Anda berhasil diperbarui! Silakan login menggunakan password baru.', 'success')
            return redirect(url_for('admin.login'))
            
    return render_template('admin/reset_password.html', token=token)


# ─── LOGOUT ──────────────────────────────────────────
@admin_bp.route('/logout')
@login_required
def logout():
    log_aktivitas('logout', 'melakukan logout keluar dari sistem')
    logout_user()
    return redirect(url_for('admin.login'))


# ─── DASHBOARD (FIX VARIABEL KATEGORI_STATS & LOGS) ──
@admin_bp.route('/')
@login_required
def dashboard():
    total_lokasi   = Lokasi.query.count()
    total_kategori = Kategori.query.count()
    kategori_stats = Kategori.query.all()
    lokasi_terbaru = Lokasi.query.order_by(Lokasi.created_at.desc()).limit(5).all()
    
    # Ambil 10 audit logs terbaru untuk ditampilkan di bento grid
    logs = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(10).all()
    
    return render_template('admin/dashboard.html',
                           total_lokasi=total_lokasi,
                           total_kategori=total_kategori,
                           kategori_stats=kategori_stats,
                           lokasi_terbaru=lokasi_terbaru,
                           logs=logs)


# ─── LIST KATEGORI ────────────────────────────────────
@admin_bp.route('/kategori')
@login_required
def kategori_list():
    kategori = Kategori.query.all()
    return render_template('admin/kategori_list.html', kategori=kategori)


# ─── TAMBAH KATEGORI ──────────────────────────────────
@admin_bp.route('/kategori/tambah', methods=['GET', 'POST'])
@login_required
def kategori_tambah():
    if request.method == 'POST':
        nama  = request.form.get('nama_kategori')
        icon  = request.form.get('icon')
        warna = request.form.get('warna')

        kategori_baru = Kategori(nama_kategori=nama, icon=icon, warna=warna)
        db.session.add(kategori_baru)
        db.session.commit()
        
        log_aktivitas('tambah', f'menambahkan kategori baru "{nama}"')
        
        flash('Kategori berhasil ditambahkan.', 'success')
        return redirect(url_for('admin.kategori_list'))

    return render_template('admin/kategori_list.html')


# ─── EDIT KATEGORI ────────────────────────────────────
@admin_bp.route('/kategori/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def kategori_edit(id):
    kategori = Kategori.query.get_or_404(id)

    if request.method == 'POST':
        kategori.nama_kategori = request.form.get('nama_kategori')
        kategori.icon          = request.form.get('icon')
        kategori.warna         = request.form.get('warna')
        db.session.commit()
        
        log_aktivitas('edit', f'memperbarui data konfigurasi kategori "{kategori.nama_kategori}"')
        
        flash('Kategori berhasil diperbarui.', 'success')
        return redirect(url_for('admin.kategori_list'))

    return render_template('admin/kategori_list.html', kategori_edit=kategori)


# ─── HAPUS KATEGORI ───────────────────────────────────
@admin_bp.route('/kategori/hapus/<int:id>', methods=['POST'])
@login_required
def kategori_hapus(id):
    kategori = Kategori.query.get_or_404(id)
    nama_kategori = kategori.nama_kategori
    jumlah_lokasi = len(kategori.lokasi)

    # Cegah penghapusan jika kategori masih dipakai oleh data lokasi
    # (kolom kategori_id di tabel lokasi bersifat NOT NULL, sehingga
    # tidak bisa di-set NULL secara otomatis)
    if jumlah_lokasi > 0:
        flash(
            f'Kategori "{nama_kategori}" tidak dapat dihapus karena masih digunakan oleh '
            f'{jumlah_lokasi} data lokasi. Silakan pindahkan atau hapus lokasi tersebut '
            f'terlebih dahulu melalui menu Kelola Lokasi.',
            'danger'
        )
        return redirect(url_for('admin.kategori_list'))

    db.session.delete(kategori)
    db.session.commit()
    
    log_aktivitas('hapus', f'menghapus kategori "{nama_kategori}"')
    
    flash('Kategori berhasil dihapus.', 'success')
    return redirect(url_for('admin.kategori_list'))


# ─── LIST LOKASI ─────────────────────────────────────
@admin_bp.route('/lokasi')
@login_required
def lokasi_list():
    lokasi = Lokasi.query.all()
    kategori_dict = {k.id: k.nama_kategori for k in Kategori.query.all()}
    return render_template('admin/lokasi_list.html', lokasi=lokasi, kategori_dict=kategori_dict)


# ─── TAMBAH LOKASI ────────────────────────────────────
@admin_bp.route('/lokasi/tambah', methods=['GET', 'POST'])
@login_required
def lokasi_tambah():
    kategori = Kategori.query.all()

    if request.method == 'POST':
        nama        = request.form.get('nama')
        kategori_id = request.form.get('kategori_id')
        deskripsi   = request.form.get('deskripsi')
        alamat      = request.form.get('alamat')
        latitude    = request.form.get('latitude')
        longitude   = request.form.get('longitude')
        kontak      = request.form.get('kontak')

        lokasi_baru = Lokasi(
            nama=nama,
            kategori_id=kategori_id,
            deskripsi=deskripsi,
            alamat=alamat,
            latitude=float(latitude),
            longitude=float(longitude),
            kontak=kontak,
            foto=None 
        )
        db.session.add(lokasi_baru)
        db.session.commit() 

        nama_file_tersimpan = []
        if 'foto' in request.files:
            file_foto_list = request.files.getlist('foto')
            
            if file_foto_list and any(f.filename != '' for f in file_foto_list):
                folder_lokasi_path = os.path.join(current_app.root_path, 'static', 'uploads', f'lokasi_{lokasi_baru.id}')
                
                if not os.path.exists(folder_lokasi_path):
                    os.makedirs(folder_lokasi_path)

                for file in file_foto_list:
                    if file and file.filename != '' and allowed_file(file.filename):
                        foto_filename = secure_filename(file.filename)
                        file.save(os.path.join(folder_lokasi_path, foto_filename))
                        nama_file_tersimpan.append(foto_filename)

        if nama_file_tersimpan:
            lokasi_baru.foto = ",".join(nama_file_tersimpan)
            db.session.commit()

        log_aktivitas('tambah', f'menambahkan objek spasial lokasi baru "{nama}"')

        flash('Lokasi baru beserta folder dokumentasi berhasil disimpan.', 'success')
        return redirect(url_for('admin.lokasi_list'))

    return render_template('admin/lokasi_form.html', kategori=kategori, lokasi=None)


# ─── EDIT LOKASI ─────────────────────────────────────
@admin_bp.route('/lokasi/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def lokasi_edit(id):
    lokasi   = Lokasi.query.get_or_404(id)
    kategori = Kategori.query.all()

    if request.method == 'POST':
        lokasi.nama        = request.form.get('nama')
        lokasi.kategori_id = request.form.get('kategori_id')
        lokasi.deskripsi   = request.form.get('deskripsi')
        lokasi.alamat      = request.form.get('alamat')
        lokasi.latitude    = float(request.form.get('latitude'))
        lokasi.longitude   = float(request.form.get('longitude'))
        lokasi.kontak      = request.form.get('kontak')

        if 'foto' in request.files:
            file_foto_list = request.files.getlist('foto')
            
            if file_foto_list and any(f.filename != '' for f in file_foto_list):
                folder_lokasi_path = os.path.join(current_app.root_path, 'static', 'uploads', f'lokasi_{lokasi.id}')
                
                if not os.path.exists(folder_lokasi_path):
                    os.makedirs(folder_lokasi_path)

                foto_saat_ini = lokasi.foto.split(',') if lokasi.foto else []
                foto_saat_ini = [f.strip() for f in foto_saat_ini if f.strip() != '']

                for file in file_foto_list:
                    if file and file.filename != '' and allowed_file(file.filename):
                        foto_filename = secure_filename(file.filename)
                        file.save(os.path.join(folder_lokasi_path, foto_filename))
                        
                        if foto_filename not in foto_saat_ini:
                            foto_saat_ini.append(foto_filename)

                lokasi.foto = ",".join(foto_saat_ini)

        db.session.commit()
        
        log_aktivitas('edit', f'memperbarui data spasial dan koordinat lokasi "{lokasi.nama}"')
        
        flash('Lokasi dan folder dokumentasi foto berhasil diperbarui.', 'success')
        return redirect(url_for('admin.lokasi_list'))

    return render_template('admin/lokasi_form.html', kategori=kategori, lokasi=lokasi)


# ─── HAPUS LOKASI ────────────────────────────────────
@admin_bp.route('/lokasi/hapus/<int:id>', methods=['POST'])
@login_required
def lokasi_hapus(id):
    lokasi = Lokasi.query.get_or_404(id)
    nama_lokasi = lokasi.nama
    
    folder_lokasi_path = os.path.join(current_app.root_path, 'static', 'uploads', f'lokasi_{lokasi.id}')
    if os.path.exists(folder_lokasi_path):
        shutil.rmtree(folder_lokasi_path) 

    db.session.delete(lokasi)
    db.session.commit()
    
    log_aktivitas('hapus', f'menghapus data lokasi "{nama_lokasi}" beserta folder penyimpanannya')
    
    flash('Lokasi dan folder dokumen terkait berhasil dihapus permanent.', 'success')
    return redirect(url_for('admin.lokasi_list'))


# ─── HAPUS MASSAL (BULK DELETE) ──────────────────────
@admin_bp.route('/lokasi/hapus-massal', methods=['POST'])
@login_required
def lokasi_hapus_massal():
    ids_terpilih = request.form.getlist('lokasi_ids')
    
    if not ids_terpilih:
        flash('Tidak ada lokasi yang dipilih untuk dihapus.', 'warning')
        return redirect(url_for('admin.lokasi_list'))
    
    try:
        jumlah_terhapus = 0
        for id_lokasi in ids_terpilih:
            lokasi_data = Lokasi.query.get(id_lokasi)
            if lokasi_data:
                folder_lokasi_path = os.path.join(current_app.root_path, 'static', 'uploads', f'lokasi_{lokasi_data.id}')
                if os.path.exists(folder_lokasi_path):
                    shutil.rmtree(folder_lokasi_path)
                
                db.session.delete(lokasi_data)
                jumlah_terhapus += 1
                
        db.session.commit()
        
        log_aktivitas('hapus', f'melakukan pembersihan massal, menghapus {jumlah_terhapus} lokasi sekaligus')
        flash(f'Berhasil menghapus massal {jumlah_terhapus} data lokasi beserta berkas foldernya.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal melakukan hapus massal: {str(e)}', 'danger')
        
    return redirect(url_for('admin.lokasi_list'))


# ─── GANTI PASSWORD ───────────────────────────────────
@admin_bp.route('/ganti-password', methods=['GET', 'POST'])
@login_required
def ganti_password():
    if request.method == 'POST':
        password_lama  = request.form.get('password_lama')
        password_baru  = request.form.get('password_baru')
        konfirmasi     = request.form.get('konfirmasi')

        if not check_password_hash(current_user.password, password_lama):
            flash('Password lama tidak sesuai.', 'danger')
        elif password_baru != konfirmasi:
            flash('Konfirmasi password tidak cocok.', 'danger')
        elif len(password_baru) < 6:
            flash('Password baru minimal 6 karakter.', 'danger')
        else:
            current_user.password = generate_password_hash(password_baru)
            db.session.commit()
            
            log_aktivitas('edit', 'berhasil memperbarui kata sandi akun admin')
            flash('Password berhasil diubah.', 'success')

    return render_template('admin/ganti_password.html')


# ─── EXPORT EXCEL ─────────────────────────────────────
@admin_bp.route('/lokasi/export')
@login_required
def lokasi_export():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Lokasi"

    header_fill = PatternFill(start_color="1a6b3c", end_color="1a6b3c", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ['No', 'Nama', 'Kategori', 'Deskripsi', 'Alamat', 'Latitude', 'Longitude', 'Kontak']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15

    lokasi_list = Lokasi.query.all()
    for i, l in enumerate(lokasi_list, 1):
        ws.append([
            i, l.nama, l.kategori.nama_kategori, l.deskripsi or '',
            l.alamat or '', l.latitude, l.longitude, l.kontak or ''
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_aktivitas('info', 'mengunduh/export data rekapitulasi seluruh lokasi ke Excel')

    return send_file(
        buffer, as_attachment=True,
        download_name='data_lokasi_karang_harapan.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ─── TEMPLATE IMPORT EXCEL ────────────────────────────
@admin_bp.route('/lokasi/template')
@login_required
def lokasi_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Import"

    header_fill = PatternFill(start_color="1a6b3c", end_color="1a6b3c", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ['nama', 'kategori', 'deskripsi', 'alamat', 'latitude', 'longitude', 'kontak']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15

    kategori_list = Kategori.query.all()
    nama_kategori = kategori_list[0].nama_kategori if kategori_list else 'UMKM'

    ws.append(['Contoh: Warung Bu Ani', nama_kategori, 'Deskripsi lokasi', 'RT 03 Karang Harapan', 3.3021, 117.5785, '08123456789'])

    ws2 = wb.create_sheet("Petunjuk")
    ws2['A1'] = "PETUNUJK PENGISIAN"
    ws2['A1'].font = Font(bold=True, size=13)
    ws2['A3'] = "Kolom yang WAJIB diisi:"
    ws2['A4'] = "- nama        : Nama lokasi"
    ws2['A5'] = "- kategori    : Harus sesuai dengan nama kategori yang terdaftar"
    ws2['A6'] = "- latitude    : Koordinat lintang (contoh: 3.3021)"
    ws2['A7'] = "- longitude   : Koordinat bujur (contoh: 117.5785)"
    ws2['A9'] = "Kolom OPSIONAL:"
    ws2['A10'] = "- deskripsi, alamat, kontak"
    ws2['A12'] = "Kategori yang tersedia:"
    for i, k in enumerate(kategori_list):
        ws2.cell(row=13+i, column=1, value=f"- {k.nama_kategori}")

    ws2.column_dimensions['A'].width = 50

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer, as_attachment=True,
        download_name='template_import_lokasi.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ─── IMPORT EXCEL ─────────────────────────────────────
@admin_bp.route('/lokasi/import', methods=['GET', 'POST'])
@login_required
def lokasi_import():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Tidak ada file yang dipilih.', 'danger')
            return redirect(url_for('admin.lokasi_import'))

        file = request.files['file']
        if file.filename == '':
            flash('Tidak ada file yang dipilih.', 'danger')
            return redirect(url_for('admin.lokasi_import'))

        try:
            wb = load_workbook(file)
            ws = wb.active

            kategori_dict = {k.nama_kategori.lower(): k.id for k in Kategori.query.all()}
            berhasil, gagal, pesan_gagal = 0, 0, []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[0]:
                    continue

                nama      = str(row[0]).strip() if row[0] else None
                kategori  = str(row[1]).strip().lower() if row[1] else None
                deskripsi = str(row[2]).strip() if row[2] else None
                alamat    = str(row[3]).strip() if row[3] else None
                latitude  = row[4]
                longitude = row[5]
                kontak    = str(row[6]).strip() if row[6] else None

                if not nama:
                    gagal += 1
                    pesan_gagal.append(f"Baris {row_num}: Nama kosong.")
                    continue

                if not latitude or not longitude:
                    gagal += 1
                    pesan_gagal.append(f"Baris {row_num}: Koordinat kosong.")
                    continue

                kategori_id = kategori_dict.get(kategori)
                if not kategori_id:
                    gagal += 1
                    pesan_gagal.append(f"Baris {row_num}: Kategori '{row[1]}' tidak ditemukan.")
                    continue

                lokasi_baru = Lokasi(
                    nama=nama, kategori_id=kategori_id, deskripsi=deskripsi,
                    alamat=alamat, latitude=float(latitude), longitude=float(longitude), kontak=kontak
                )
                db.session.add(lokasi_baru)
                berhasil += 1

            db.session.commit()

            if gagal == 0:
                log_aktivitas('tambah', f'berhasil melakukan import massal {berhasil} lokasi via Excel')
                flash(f'{berhasil} lokasi berhasil diimport.', 'success')
            else:
                log_aktivitas('tambah', f'melakukan import lokasi via Excel ({berhasil} sukses, {gagal} gagal)')
                flash(f'{berhasil} lokasi berhasil, {gagal} gagal: ' + ' | '.join(pesan_gagal), 'warning')

        except Exception as e:
            flash(f'Gagal membaca file: {str(e)}', 'danger')

        return redirect(url_for('admin.lokasi_list'))

    return render_template('admin/lokasi_import.html')